#!/usr/bin/env python3
"""
TellyAds Editorial Import Script
Import Wix/Excel data into ad_editorial table

Usage:
  python scripts/import_editorial.py --input data/wix.xlsx --dry-run
  python scripts/import_editorial.py --input data/wix.xlsx --apply
  python scripts/import_editorial.py --input data/wix.xlsx --apply --force-overwrite
"""

import argparse
import csv
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, List, Tuple
from urllib.parse import urlparse

# Load .env file
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # dotenv not installed, use env vars directly

# Support both xlsx and csv
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

import psycopg2
from psycopg2.extras import RealDictCursor

# --------------------------------------------------------------------------
# Configuration
# --------------------------------------------------------------------------

# Column name mappings (flexible to handle different Excel/CSV exports)
# Keys are canonical names, values are possible column headers (case-insensitive)
COLUMN_ALIASES = {
    'legacy_url': ['legacy_url', 'url', 'wix_url', 'original_url', 'link',
                   'telly ads (advertiser-1, commercial...)',  # Wix export format
                   'telly ads (advertiser-1, commerci...)'],
    'title': ['title', 'headline', 'name', 'ad_title', 'editorial_title',
              'commercial_title'],  # Wix export
    'description': ['description', 'summary', 'editorial_summary', 'body', 'content'],
    'brand': ['brand', 'brand_name', 'advertiser', 'company',
              'advertiser-1'],  # Wix export
    'year': ['year', 'ad_year', 'release_year'],
    'external_id': ['external_id', 'ta_id', 'tellyads_id', 'id',
                    'movie_filename'],  # Wix export - "TA8887"
    'wix_item_id': ['wix_item_id', 'wix_id', 'cms_id', '_id',
                    'record_id'],  # Wix export
    'publish_status': ['publish_status', 'status', 'state'],
    'publish_date': ['publish_date', 'published_at', 'date_published', 'original_publish_date',
                     'date_collected'],  # Wix export - DD/MM/YYYY format
    'curated_tags': ['curated_tags', 'tags', 'categories'],
    'duration': ['duration', 'length'],  # Wix export - seconds
    'views': ['views', 'view_count'],  # Wix export - for initializing feedback
    'thumbnail_url': ['thumbnail_url', 'still_filename_link'],  # Wix export
    'video_url': ['video_url', 'vid_filename_link'],  # Wix export
}

# --------------------------------------------------------------------------
# Slugify Helper
# --------------------------------------------------------------------------

def slugify(text: str) -> str:
    """Convert text to URL-safe slug"""
    if not text:
        return ''
    # Lowercase
    slug = text.lower()
    # Replace special chars with hyphens
    slug = re.sub(r'[^\w\s-]', '', slug)
    # Replace whitespace with hyphens
    slug = re.sub(r'[-\s]+', '-', slug)
    # Strip leading/trailing hyphens
    slug = slug.strip('-')
    return slug


def extract_slug_from_url(url: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Extract brand_slug and slug from legacy Wix URL

    Patterns supported:
    - /ads/{brand}/{slug}
    - /adverts/{brand}/{slug}
    - /ad-detail/{brand}/{slug}
    - /collection/{brand}/{slug}
    - /{brand}/{slug} (2-segment paths)
    - Full URLs with paths

    Returns: (brand_slug, slug) or (None, None) if cannot parse
    """
    if not url:
        return None, None

    # Parse URL
    parsed = urlparse(url)
    path = parsed.path.strip('/')

    if not path:
        return None, None

    segments = path.split('/')

    # Remove known prefixes
    prefixes_to_strip = ['ads', 'adverts', 'ad-detail', 'advert', 'collection', 'item']
    while segments and segments[0].lower() in prefixes_to_strip:
        segments = segments[1:]

    # Need at least 2 segments: brand/slug
    if len(segments) >= 2:
        brand_slug = slugify(segments[0])
        slug = slugify(segments[1])
        return brand_slug, slug
    elif len(segments) == 1:
        # Only slug, no brand - will need manual matching
        return None, slugify(segments[0])

    return None, None


# --------------------------------------------------------------------------
# Data Loading
# --------------------------------------------------------------------------

def load_excel(file_path: str) -> List[Dict]:
    """Load data from Excel file"""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required for Excel files: pip install openpyxl")

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    # Get headers from first row
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).lower().strip() if cell.value else '')

    # Map headers to canonical names
    header_map = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        for i, h in enumerate(headers):
            if h in aliases:
                header_map[i] = canonical
                break

    # Load rows
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for i, value in enumerate(row):
            if i in header_map:
                record[header_map[i]] = value
        if any(record.values()):  # Skip empty rows
            rows.append(record)

    return rows


def load_csv(file_path: str) -> List[Dict]:
    """Load data from CSV file"""
    rows = []
    with open(file_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)

        # Map headers to canonical names
        header_map = {}
        for canonical, aliases in COLUMN_ALIASES.items():
            for h in reader.fieldnames or []:
                if h.lower().strip() in aliases:
                    header_map[h] = canonical
                    break

        for row in reader:
            record = {}
            for orig_name, value in row.items():
                if orig_name in header_map:
                    record[header_map[orig_name]] = value
            if any(record.values()):
                rows.append(record)

    return rows


def load_data(file_path: str) -> List[Dict]:
    """Load data from Excel or CSV"""
    ext = Path(file_path).suffix.lower()
    if ext in ['.xlsx', '.xls']:
        return load_excel(file_path)
    elif ext == '.csv':
        return load_csv(file_path)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


# --------------------------------------------------------------------------
# Matching Logic (Optimized - load all ads upfront)
# --------------------------------------------------------------------------

def load_ads_cache(conn) -> Tuple[Dict, Dict, Dict]:
    """
    Load all ads and editorial records upfront for fast in-memory matching.
    Returns:
        - ads_by_external_id: {external_id: ad_id}
        - editorial_by_wix_id: {wix_item_id: ad_id}
        - ads_by_brand_year: {(brand_lower, year): [(ad_id, title, external_id), ...]}
    """
    print("Loading ads cache from database...", flush=True)

    ads_by_external_id = {}
    ads_by_brand_year = {}
    editorial_by_wix_id = {}

    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        # Load all ads
        cur.execute("""
            SELECT id, external_id, brand_name, one_line_summary, year
            FROM ads
        """)
        ads = cur.fetchall()
        print(f"  Loaded {len(ads)} ads", flush=True)

        for ad in ads:
            ad_id = str(ad['id'])
            ext_id = ad['external_id']
            if ext_id:
                ads_by_external_id[ext_id.upper()] = ad_id

            # Index by brand + year
            brand = (ad['brand_name'] or '').lower()
            year = ad['year']
            key = (brand, year)
            if key not in ads_by_brand_year:
                ads_by_brand_year[key] = []
            ads_by_brand_year[key].append((ad_id, ad['one_line_summary'], ext_id))

        # Load existing editorial records
        cur.execute("SELECT ad_id, wix_item_id FROM ad_editorial WHERE wix_item_id IS NOT NULL")
        editorial = cur.fetchall()
        print(f"  Loaded {len(editorial)} existing editorial records", flush=True)

        for e in editorial:
            if e['wix_item_id']:
                editorial_by_wix_id[e['wix_item_id']] = str(e['ad_id'])

    return ads_by_external_id, editorial_by_wix_id, ads_by_brand_year


def match_to_ad_cached(record: Dict, ads_by_external_id: Dict,
                       editorial_by_wix_id: Dict, ads_by_brand_year: Dict) -> Tuple[Optional[str], str]:
    """
    Match import record to existing ad using in-memory cache.

    Priority:
    1. external_id exact match
    2. wix_item_id match (if already imported)
    3. brand + year (fuzzy)
    4. Manual review bucket

    Returns: (ad_id, match_method) or (None, 'unmatched')
    """
    # Priority 1: external_id
    external_id = record.get('external_id')
    if external_id:
        ext_str = str(external_id).strip()
        if not ext_str.upper().startswith('TA'):
            ext_str = f"TA{ext_str}"

        ad_id = ads_by_external_id.get(ext_str.upper())
        if ad_id:
            return ad_id, 'external_id'

    # Priority 2: wix_item_id (check if already imported)
    wix_id = record.get('wix_item_id')
    if wix_id:
        ad_id = editorial_by_wix_id.get(wix_id)
        if ad_id:
            return ad_id, 'wix_item_id_existing'

    # Priority 3: Brand + Year fuzzy match
    # DISABLED: This creates false positives when only one ad per brand exists
    # TODO: Re-enable with stricter title matching when more ads are indexed
    # brand = record.get('brand')
    # year = record.get('year')
    # if brand:
    #     year_val = None
    #     if year:
    #         try:
    #             year_val = int(year)
    #         except (ValueError, TypeError):
    #             pass
    #     key = (brand.lower(), year_val)
    #     candidates = ads_by_brand_year.get(key, [])
    #     if len(candidates) == 1:
    #         return candidates[0][0], 'brand_year_unique'
    #     elif len(candidates) > 1:
    #         return None, 'multiple_matches'

    return None, 'unmatched'


# --------------------------------------------------------------------------
# Import Logic
# --------------------------------------------------------------------------

def import_record(conn, record: Dict, ad_id: str, force_overwrite: bool = False) -> Dict:
    """
    Import single record into ad_editorial

    Returns result dict with status
    """
    # Parse legacy URL for slug
    legacy_url = record.get('legacy_url', '')
    brand_slug_from_url, slug_from_url = extract_slug_from_url(legacy_url)

    # Build brand_slug
    brand = record.get('brand', '')
    brand_slug = brand_slug_from_url or slugify(brand)

    if not brand_slug:
        return {'status': 'error', 'reason': 'Cannot determine brand_slug'}

    # Build slug
    title = record.get('title', '')
    slug = slug_from_url or slugify(title)

    if not slug:
        return {'status': 'error', 'reason': 'Cannot determine slug'}

    # Parse other fields
    year = record.get('year')
    if year:
        try:
            year = int(year)
        except (ValueError, TypeError):
            year = None

    publish_status = record.get('publish_status', 'draft')
    if publish_status not in ('draft', 'published', 'archived'):
        publish_status = 'draft'

    publish_date = record.get('publish_date')
    if publish_date and isinstance(publish_date, str):
        publish_date = publish_date.strip()
        parsed_date = None
        # Try DD/MM/YYYY format (Wix export)
        if '/' in publish_date:
            try:
                parsed_date = datetime.strptime(publish_date, '%d/%m/%Y')
            except ValueError:
                try:
                    parsed_date = datetime.strptime(publish_date, '%m/%d/%Y')
                except ValueError:
                    pass
        # Try ISO format
        if not parsed_date:
            try:
                parsed_date = datetime.fromisoformat(publish_date.replace('Z', '+00:00'))
            except ValueError:
                pass
        publish_date = parsed_date
    elif publish_date and isinstance(publish_date, datetime):
        pass  # Already a datetime
    else:
        publish_date = None

    curated_tags = record.get('curated_tags', '')
    if isinstance(curated_tags, str):
        curated_tags = [t.strip() for t in curated_tags.split(',') if t.strip()]
    elif not curated_tags:
        curated_tags = []

    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        # Check for existing editorial record
        cur.execute("""
            SELECT id, brand_slug, slug, headline, editorial_summary
            FROM ad_editorial
            WHERE ad_id = %s
        """, (ad_id,))
        existing = cur.fetchone()

        if existing:
            if not force_overwrite:
                # Don't overwrite existing human content
                return {
                    'status': 'skipped',
                    'reason': 'Editorial exists (use --force-overwrite)',
                    'existing_slug': f"{existing['brand_slug']}/{existing['slug']}"
                }

            # Update existing record
            cur.execute("""
                UPDATE ad_editorial SET
                    brand_slug = COALESCE(%s, brand_slug),
                    slug = COALESCE(%s, slug),
                    headline = COALESCE(%s, headline),
                    editorial_summary = COALESCE(%s, editorial_summary),
                    curated_tags = CASE WHEN %s::text[] != '{}' THEN %s ELSE curated_tags END,
                    wix_item_id = COALESCE(%s, wix_item_id),
                    original_publish_date = COALESCE(%s, original_publish_date),
                    override_year = COALESCE(%s, override_year),
                    legacy_url = COALESCE(%s, legacy_url),
                    status = %s,
                    updated_at = now()
                WHERE ad_id = %s
                RETURNING id
            """, (
                brand_slug, slug,
                record.get('title'), record.get('description'),
                curated_tags, curated_tags,
                record.get('wix_item_id'), publish_date, year,
                legacy_url if legacy_url else None,
                publish_status,
                ad_id
            ))
            return {'status': 'updated', 'slug': f"{brand_slug}/{slug}"}

        else:
            # Check for slug conflict
            cur.execute("""
                SELECT id FROM ad_editorial
                WHERE brand_slug = %s AND slug = %s
            """, (brand_slug, slug))
            if cur.fetchone():
                # Append suffix to make unique
                base_slug = slug
                for i in range(2, 100):
                    test_slug = f"{base_slug}-{i}"
                    cur.execute("""
                        SELECT id FROM ad_editorial
                        WHERE brand_slug = %s AND slug = %s
                    """, (brand_slug, test_slug))
                    if not cur.fetchone():
                        slug = test_slug
                        break
                else:
                    return {'status': 'error', 'reason': f'Slug conflict: {brand_slug}/{base_slug}'}

            # Insert new record
            cur.execute("""
                INSERT INTO ad_editorial (
                    ad_id, brand_slug, slug,
                    headline, editorial_summary, curated_tags,
                    wix_item_id, original_publish_date,
                    override_year, legacy_url, status
                ) VALUES (
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s,
                    %s, %s, %s
                )
                RETURNING id
            """, (
                ad_id, brand_slug, slug,
                record.get('title'), record.get('description'), curated_tags,
                record.get('wix_item_id'), publish_date,
                year, legacy_url if legacy_url else None, publish_status
            ))
            return {'status': 'inserted', 'slug': f"{brand_slug}/{slug}"}


# --------------------------------------------------------------------------
# Main Import Flow
# --------------------------------------------------------------------------

def run_import(file_path: str, db_url: str, dry_run: bool = True, force_overwrite: bool = False):
    """
    Main import function
    """
    print(f"Loading data from: {file_path}")
    records = load_data(file_path)
    print(f"Loaded {len(records)} records")

    conn = psycopg2.connect(db_url)

    # Load all ads upfront for fast in-memory matching
    ads_by_external_id, editorial_by_wix_id, ads_by_brand_year = load_ads_cache(conn)

    results = {
        'matched': [],
        'unmatched': [],
        'conflicts': [],
        'inserted': [],
        'updated': [],
        'skipped': [],
        'errors': [],
    }

    total = len(records)
    for i, record in enumerate(records):
        row_num = i + 2  # Excel row number (1-indexed + header)

        # Progress every 1000 records
        if i % 1000 == 0:
            print(f"Processing {i}/{total} ({i*100//total}%)...", flush=True)

        # Match to ad using in-memory cache
        ad_id, match_method = match_to_ad_cached(
            record, ads_by_external_id, editorial_by_wix_id, ads_by_brand_year
        )

        record_info = {
            'row': row_num,
            'legacy_url': record.get('legacy_url', ''),
            'title': record.get('title', ''),
            'brand': record.get('brand', ''),
            'match_method': match_method,
        }

        if not ad_id:
            if match_method == 'multiple_matches':
                results['conflicts'].append(record_info)
            else:
                results['unmatched'].append(record_info)
            continue

        record_info['ad_id'] = ad_id
        results['matched'].append(record_info)

        if not dry_run:
            # Actually import
            import_result = import_record(conn, record, ad_id, force_overwrite)
            record_info.update(import_result)

            if import_result['status'] == 'inserted':
                results['inserted'].append(record_info)
            elif import_result['status'] == 'updated':
                results['updated'].append(record_info)
            elif import_result['status'] == 'skipped':
                results['skipped'].append(record_info)
            elif import_result['status'] == 'error':
                results['errors'].append(record_info)

    if not dry_run:
        conn.commit()
    conn.close()

    # Output results
    output_dir = Path(file_path).parent
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    # Write CSVs
    for name in ['matched', 'unmatched', 'conflicts']:
        if results[name]:
            out_file = output_dir / f"{name}_{timestamp}.csv"
            with open(out_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=results[name][0].keys())
                writer.writeheader()
                writer.writerows(results[name])
            print(f"Wrote: {out_file}")

    # Write summary JSON
    summary = {
        'timestamp': timestamp,
        'input_file': str(file_path),
        'dry_run': dry_run,
        'force_overwrite': force_overwrite,
        'total_records': len(records),
        'matched_count': len(results['matched']),
        'unmatched_count': len(results['unmatched']),
        'conflicts_count': len(results['conflicts']),
        'inserted_count': len(results['inserted']),
        'updated_count': len(results['updated']),
        'skipped_count': len(results['skipped']),
        'error_count': len(results['errors']),
        'match_rate': len(results['matched']) / len(records) * 100 if records else 0,
    }

    summary_file = output_dir / f"import_summary_{timestamp}.json"
    with open(summary_file, 'w') as f:
        json.dump(summary, f, indent=2)
    print(f"Wrote: {summary_file}")

    # Console summary
    print("\n" + "="*60)
    print("IMPORT SUMMARY")
    print("="*60)
    print(f"Total records:     {summary['total_records']}")
    print(f"Matched:           {summary['matched_count']} ({summary['match_rate']:.1f}%)")
    print(f"Unmatched:         {summary['unmatched_count']}")
    print(f"Conflicts:         {summary['conflicts_count']}")
    if not dry_run:
        print(f"Inserted:          {summary['inserted_count']}")
        print(f"Updated:           {summary['updated_count']}")
        print(f"Skipped:           {summary['skipped_count']}")
        print(f"Errors:            {summary['error_count']}")
    else:
        print("\n[DRY RUN - No changes made]")
    print("="*60)

    return summary


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description='Import Wix/Excel data into TellyAds editorial')
    parser.add_argument('--input', required=True, help='Path to Excel/CSV file')
    parser.add_argument('--dry-run', action='store_true', help='Preview without writing')
    parser.add_argument('--apply', action='store_true', help='Actually apply changes')
    parser.add_argument('--force-overwrite', action='store_true', help='Overwrite existing editorial')
    parser.add_argument('--db-url', default=None, help='Database URL (default: SUPABASE_DB_URL env)')

    args = parser.parse_args()

    if not args.apply and not args.dry_run:
        args.dry_run = True
        print("Note: Running in dry-run mode (use --apply to write changes)")

    db_url = args.db_url or os.environ.get('SUPABASE_DB_URL')
    if not db_url:
        print("Error: Database URL required (--db-url or SUPABASE_DB_URL env)")
        sys.exit(1)

    run_import(
        file_path=args.input,
        db_url=db_url,
        dry_run=not args.apply,
        force_overwrite=args.force_overwrite
    )


if __name__ == '__main__':
    main()
